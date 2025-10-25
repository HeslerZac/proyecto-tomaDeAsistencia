import os
import threading
import time
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from typing import Dict, Optional, Tuple
from sqlalchemy import or_

from flask import Flask, jsonify, redirect, render_template, request, url_for, flash, make_response
import csv
import json
from io import StringIO
from flask_sqlalchemy import SQLAlchemy

from config import Config

try:
    import serial  # type: ignore
    import serial.tools.list_ports  # type: ignore
except Exception as e:  # pragma: no cover
    serial = None


db = SQLAlchemy()
try:
    from flask_wtf import CSRFProtect
except Exception:
    CSRFProtect = None  # type: ignore

try:
    from flask_login import (
        LoginManager,
        login_user,
        logout_user,
        login_required,
        current_user,
        UserMixin,
    )
except Exception:
    LoginManager = None  # type: ignore
    login_user = logout_user = login_required = lambda *a, **k: (lambda f: f)  # type: ignore
    current_user = None  # type: ignore
    UserMixin = object  # type: ignore


class Persona(db.Model):
    __tablename__ = "persona"
    id = db.Column(db.Integer, primary_key=True)  # Sensor fingerID
    nombre = db.Column(db.Text, nullable=False)
    rol = db.Column(db.Text, nullable=False)  # 'ALUMNO' | 'PROFESOR' | 'PERSONAL'
    huella_guardada = db.Column(db.Boolean, default=False, nullable=False)
    grado = db.Column(db.Text)  # opcional: grado/área/año
    seccion = db.Column(db.Text)  # opcional: sección/turno
    documento = db.Column(db.Text)  # DNI/código
    activo = db.Column(db.Boolean, default=True, nullable=False)

    def as_dict(self):
        return {
            "id": self.id,
            "nombre": self.nombre,
            "rol": self.rol,
            "huella_guardada": self.huella_guardada,
            "grado": self.grado,
            "seccion": self.seccion,
            "documento": self.documento,
            "activo": self.activo,
        }


class Asistencia(db.Model):
    __tablename__ = "asistencia"
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    person_id = db.Column(db.Integer, db.ForeignKey("persona.id"), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    persona = db.relationship(Persona, backref="asistencias")

    def as_dict(self):
        return {
            "id": self.id,
            "person_id": self.person_id,
            "timestamp": self.timestamp.isoformat(),
        }


class User(db.Model, UserMixin):
    __tablename__ = "user"
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    username = db.Column(db.String(64), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    is_admin = db.Column(db.Boolean, default=False, nullable=False)



class SerialManager:
    def __init__(self, app: Flask):
        self.app = app
        self.port_name = app.config.get("SERIAL_PORT")
        self.baudrate = app.config.get("SERIAL_BAUDRATE", 9600)
        self.timeout = app.config.get("SERIAL_TIMEOUT", 0.2)
        self._ser = None
        self._thread: Optional[threading.Thread] = None
        self._stop = threading.Event()
        # pending REG responses: id -> (event, result_str)
        self._pending_reg: Dict[int, Tuple[threading.Event, Dict[str, Optional[str]]]] = {}
        self._last_reg: Dict[int, Tuple[str, float]] = {}  # id -> (result, ts)
        self._pending_del: Dict[int, Tuple[threading.Event, Dict[str, Optional[str]]]] = {}
        self._last_del: Dict[int, Tuple[str, float]] = {}
        from collections import deque
        self._log = deque(maxlen=200)  # recent serial lines
        self._current_port: Optional[str] = None
        self._last_fnd: Dict[int, datetime] = {}
        self._lock = threading.Lock()

    def _normalize_port(self, name: str) -> str:
        # Windows COM ports > 9 sometimes need the \\.\ prefix for some stacks.
        try:
            up = name.upper()
            if up.startswith("COM") and up[3:].isdigit() and int(up[3:]) >= 10:
                return r"\\.\%s" % name
        except Exception:
            pass
        return name

    def _scan_candidates(self):
        cands = []
        try:
            ports = list(serial.tools.list_ports.comports()) if serial else []
            for p in ports:
                desc = (p.description or "").lower()
                if any(k in desc for k in ["arduino", "ch340", "cp210", "usb-serial", "silicon labs", "prolific", "ftdi"]):
                    cands.append(p.device)
            # fallback: include all devices if none matched
            if not cands:
                cands = [p.device for p in ports]
        except Exception:
            pass
        return cands

    def open(self):
        if serial is None:
            self.app.logger.error("pyserial not installed; serial disabled")
            return
        try:
            port_cfg = (self.port_name or "").strip()
            if not port_cfg or port_cfg.upper() in {"AUTO"}:
                # try auto-detect
                for cand in self._scan_candidates():
                    try:
                        norm = self._normalize_port(cand)
                        self._ser = serial.Serial(norm, self.baudrate, timeout=self.timeout)
                        self._current_port = cand
                        self.app.logger.info(f"Serial auto-detect: {cand} @ {self.baudrate}")
                        return
                    except Exception:
                        continue
                self._ser = None
                return
            # explicit port
            norm = self._normalize_port(port_cfg)
            self._ser = serial.Serial(norm, self.baudrate, timeout=self.timeout)
            self._current_port = port_cfg
            self.app.logger.info(f"Serial abierto en {port_cfg} @ {self.baudrate}")
        except Exception as e:
            self.app.logger.error(f"No se pudo abrir el puerto {self.port_name}: {e}")
            # Fallback: if the configured port no existe, intente auto-detectar
            if isinstance(e, FileNotFoundError):
                for cand in self._scan_candidates():
                    try:
                        norm = self._normalize_port(cand)
                        self._ser = serial.Serial(norm, self.baudrate, timeout=self.timeout)
                        self._current_port = cand
                        self.app.logger.info(f"Serial fallback a {cand} @ {self.baudrate}")
                        return
                    except Exception:
                        continue
            self._ser = None

    def start(self):
        if self._thread and self._thread.is_alive():
            return
        self._stop.clear()
        self.open()
        self._thread = threading.Thread(target=self._run, name="serial-worker", daemon=True)
        self._thread.start()

    def stop(self):
        self._stop.set()
        if self._thread:
            self._thread.join(timeout=2)
        if self._ser:
            try:
                self._ser.close()
            except Exception:
                pass

    def _ensure_open(self):
        if self._ser is None or not self._ser.is_open:
            self.open()

    def send_cmd(self, line: str):
        self._ensure_open()
        if not self._ser:
            raise RuntimeError("Serial no disponible")
        data = (line.strip() + "\n").encode("utf-8")
        self._ser.write(data)
        self._append_log(f"TX> {line.strip()}")

    def wait_reg_result(self, finger_id: int, timeout: float) -> Optional[str]:
        evt = threading.Event()
        container = {"result": None}  # type: ignore
        with self._lock:
            # Fast path: if we already have a cached result (arrived early)
            if finger_id in self._last_reg:
                result, ts = self._last_reg.pop(finger_id)
                return result
            self._pending_reg[finger_id] = (evt, container)
        finished = evt.wait(timeout)
        with self._lock:
            self._pending_reg.pop(finger_id, None)
        if not finished:
            return None
        return container["result"]  # type: ignore

    def _set_reg_result(self, finger_id: int, result: str):
        with self._lock:
            entry = self._pending_reg.get(finger_id)
        if entry:
            evt, container = entry
            container["result"] = result
            evt.set()
        else:
            # No waiter yet; cache briefly
            self._last_reg[finger_id] = (result, time.time())

    def wait_del_result(self, finger_id: int, timeout: float) -> Optional[str]:
        evt = threading.Event()
        container = {"result": None}
        with self._lock:
            if finger_id in self._last_del:
                result, ts = self._last_del.pop(finger_id)
                return result
            self._pending_del[finger_id] = (evt, container)
        finished = evt.wait(timeout)
        with self._lock:
            self._pending_del.pop(finger_id, None)
        if not finished:
            return None
        return container["result"]  # type: ignore

    def _set_del_result(self, finger_id: int, result: str):
        with self._lock:
            entry = self._pending_del.get(finger_id)
        if entry:
            evt, container = entry
            container["result"] = result
            evt.set()
        else:
            self._last_del[finger_id] = (result, time.time())

    def _run(self):
        buf = b""
        while not self._stop.is_set():
            try:
                self._ensure_open()
                if not self._ser:
                    time.sleep(1)
                    continue
                chunk = self._ser.read(256)
                if not chunk:
                    continue
                buf += chunk
                while b"\n" in buf:
                    line, buf = buf.split(b"\n", 1)
                    try:
                        text = line.decode("utf-8", errors="ignore").strip()
                    except Exception:
                        continue
                    if text:
                        self._handle_line(text)
            except Exception as e:
                self.app.logger.error(f"Error en serial worker: {e}")
                # Intentar reabrir tras un breve descanso
                try:
                    if self._ser:
                        self._ser.close()
                except Exception:
                    pass
                self._ser = None
                time.sleep(2)

    def _append_log(self, text: str):
        try:
            ts = datetime.now().strftime("%H:%M:%S")
            with self._lock:
                self._log.appendleft(f"{ts} | {text}")
        except Exception:
            pass

    def get_log(self, limit: int = 100):
        with self._lock:
            return list(list(self._log)[:limit])

    def _handle_line(self, text: str):
        self.app.logger.info(f"Serial RX: {text}")
        self._append_log(f"RX> {text}")
        # Expected messages:
        # REG_OK,<ID>
        # REG_FAIL
        # DEL_OK,<ID>
        # DEL_FAIL
        # FND,<ID>
        try:
            if text.startswith("REG_OK,"):
                fid = int(text.split(",", 1)[1])
                self._set_reg_result(fid, "REG_OK")
            elif text == "REG_FAIL":
                # If REG_FAIL doesn't include ID, release all pending
                with self._lock:
                    keys = list(self._pending_reg.keys())
                for fid in keys:
                    self._set_reg_result(fid, "REG_FAIL")
            elif text.startswith("DEL_OK,"):
                fid = int(text.split(",", 1)[1])
                self._set_del_result(fid, "DEL_OK")
            elif text == "DEL_FAIL":
                with self._lock:
                    keys = list(self._pending_del.keys())
                for fid in keys:
                    self._set_del_result(fid, "DEL_FAIL")
            elif text.startswith("FND,"):
                fid = int(text.split(",", 1)[1])
                # Register asistencia in DB
                with self.app.app_context():
                    persona = db.session.get(Persona, fid)
                    if persona is not None:
                        if getattr(persona, 'activo', True):
                            # Anti-rebote: si hay asistencia reciente para el mismo ID, ignorar
                            now_dt = datetime.now(self.app.tz) if self.app.config.get("TIME_USE_LOCAL", True) else datetime.utcnow().replace(tzinfo=timezone.utc).astimezone(self.app.tz)
                            last_dt = self._last_fnd.get(persona.id)
                            window = float(self.app.config.get("DUPLICATE_WINDOW_SEC", 60.0))
                            if last_dt is not None and (now_dt - last_dt).total_seconds() < window:
                                self._append_log(f"IGN> FND duplicado {fid} en {int((now_dt-last_dt).total_seconds())}s")
                            else:
                                # Store timestamp either in local tz (naive) or UTC (naive)
                                if self.app.config.get("TIME_USE_LOCAL", True):
                                    ts = now_dt.replace(tzinfo=None)
                                else:
                                    ts = now_dt.astimezone(timezone.utc).replace(tzinfo=None)
                                db.session.add(Asistencia(person_id=persona.id, timestamp=ts))
                                db.session.commit()
                                self._last_fnd[persona.id] = now_dt
                        else:
                            self._append_log(f"IGN> Persona {fid} inactiva; asistencia ignorada")
            # Optional: handle DEL_OK/DEL_FAIL if you add delete route
        except Exception as e:
            self.app.logger.error(f"Error procesando linea serial '{text}': {e}")


serial_manager: Optional[SerialManager] = None


def create_app() -> Flask:
    app = Flask(__name__)
    app.config.from_object(Config)
    db.init_app(app)

    # Login manager setup
    if LoginManager:
        lm = LoginManager()
        lm.init_app(app)
        lm.login_view = "login"

        @lm.user_loader
        def _load_user(user_id):  # type: ignore
            try:
                return db.session.get(User, int(user_id))
            except Exception:
                return None

    # CSRF helper for templates (safe no-op when CSRF not enabled)
    @app.context_processor
    def _inject_csrf():
        def _tok():
            try:
                from flask_wtf.csrf import generate_csrf as _gen
                return _gen()
            except Exception:
                return ""
        return {"csrf_token": _tok}

    # Timezone setup
    try:
        app.tz = ZoneInfo(app.config.get("TIMEZONE", "UTC"))  # type: ignore[attr-defined]
    except Exception:
        app.tz = ZoneInfo("UTC")  # type: ignore[attr-defined]

    @app.template_filter("localdt")
    def _localdt(dt: datetime, fmt: str = "%Y-%m-%d %H:%M:%S") -> str:
        if not isinstance(dt, datetime):
            return ""
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        try:
            return dt.astimezone(app.tz).strftime(fmt)  # type: ignore[attr-defined]
        except Exception:
            return dt.strftime(fmt)

    # Horarios y tardanzas
    def _parse_hhmm(s: str) -> Tuple[int, int]:
        try:
            h, m = s.strip().split(":", 1)
            return int(h), int(m)
        except Exception:
            return 8, 0

    app.start_time_default = app.config.get("START_TIME", "08:00")  # type: ignore[attr-defined]
    app.tardy_tol_min = int(app.config.get("TARDINESS_TOL_MIN", 10))  # type: ignore[attr-defined]
    try:
        app.schedule_by_grade = json.loads(app.config.get("SCHEDULE_BY_GRADE") or "{}")  # type: ignore[attr-defined]
    except Exception:
        app.schedule_by_grade = {}  # type: ignore[attr-defined]

    def _get_scheduled_local(persona, day_local: datetime):
        # Returns a local datetime for scheduled start
        hh, mm = _parse_hhmm(app.start_time_default)  # type: ignore[attr-defined]
        if persona and getattr(persona, "grado", None):
            g = persona.grado
            s = app.schedule_by_grade.get(g) if isinstance(app.schedule_by_grade, dict) else None  # type: ignore[attr-defined]
            if s:
                hh, mm = _parse_hhmm(str(s))
        return datetime(day_local.year, day_local.month, day_local.day, hh, mm, 0)

    def _build_report_rows(asist_list, tzinfo):
        # Build entrada/salida pairs per person per local date
        from collections import defaultdict
        bucket = defaultdict(list)  # key: (person_id, local_date) -> list[datetime]
        person_map = {}
        for a in asist_list:
            dt = a.timestamp
            if app.config.get("TIME_USE_LOCAL", True):
                # stored as local naive
                if dt.tzinfo is None:
                    loc = dt.replace(tzinfo=tzinfo)
                else:
                    loc = dt.astimezone(tzinfo)
            else:
                # stored as UTC naive
                if dt.tzinfo is None:
                    loc = dt.replace(tzinfo=timezone.utc).astimezone(tzinfo)
                else:
                    loc = dt.astimezone(tzinfo)
            key = (a.person_id, loc.date())
            bucket[key].append(loc)
            if a.persona:
                person_map[a.person_id] = a.persona
        rows = []
        for (pid, day), times in bucket.items():
            times.sort()
            for i in range(0, len(times), 2):
                entrada = times[i]
                salida = times[i + 1] if i + 1 < len(times) else None
                dur_min = int(((salida - entrada).total_seconds() // 60) if salida else 0)
                p = person_map.get(pid)
                # Calcular tardanza según horario configurado
                sched_local = _get_scheduled_local(p, datetime(day.year, day.month, day.day, 0, 0, 0))
                tol = app.tardy_tol_min  # type: ignore[attr-defined]
                tardanza = 0
                try:
                    if entrada > sched_local and ((entrada - sched_local).total_seconds() // 60) > tol:
                        tardanza = int((entrada - sched_local).total_seconds() // 60) - tol
                except Exception:
                    tardanza = 0
                rows.append({
                    "fecha": day.isoformat(),
                    "person_id": pid,
                    "nombre": p.nombre if p else None,
                    "rol": p.rol if p else None,
                    "grado": getattr(p, 'grado', None) if p else None,
                    "seccion": getattr(p, 'seccion', None) if p else None,
                    "entrada": entrada.strftime('%H:%M:%S'),
                    "salida": salida.strftime('%H:%M:%S') if salida else "",
                    "duracion_min": dur_min,
                    "tardanza_min": tardanza,
                })
        # Order by date, name, entrada
        rows.sort(key=lambda r: (r["fecha"], r["nombre"] or "", r["entrada"]))
        return rows

    with app.app_context():
        db.create_all()
        # Seed default admin user if none exists
        try:
            from werkzeug.security import generate_password_hash
            if db.session.query(User).count() == 0:
                db.session.add(User(username='admin', password_hash=generate_password_hash('admin123'), is_admin=True))
                db.session.commit()
        except Exception as _e:
            app.logger.warning(f"No se pudo crear admin por defecto: {_e}")
        # Ensure extra columns exist on Persona for existing DBs
        try:
            eng = db.get_engine()
            with eng.begin() as conn:
                conn.execute(db.text("ALTER TABLE persona ADD COLUMN IF NOT EXISTS grado TEXT"))
                conn.execute(db.text("ALTER TABLE persona ADD COLUMN IF NOT EXISTS seccion TEXT"))
                conn.execute(db.text("ALTER TABLE persona ADD COLUMN IF NOT EXISTS documento TEXT"))
                conn.execute(db.text("ALTER TABLE persona ADD COLUMN IF NOT EXISTS activo BOOLEAN DEFAULT TRUE"))
                conn.execute(db.text("CREATE UNIQUE INDEX IF NOT EXISTS persona_documento_uniq ON persona (documento) WHERE documento IS NOT NULL"))
                conn.execute(db.text("CREATE INDEX IF NOT EXISTS idx_asistencia_ts ON asistencia (timestamp DESC)"))
                conn.execute(db.text("CREATE INDEX IF NOT EXISTS idx_asistencia_person_ts ON asistencia (person_id, timestamp DESC)"))
                conn.execute(db.text("CREATE INDEX IF NOT EXISTS idx_persona_rol ON persona (rol)"))
                conn.execute(db.text("CREATE INDEX IF NOT EXISTS idx_persona_activo ON persona (activo)"))
        except Exception as _e:
            app.logger.warning(f"No se pudieron asegurar columnas extra: {_e}")

    @app.route("/")
    @login_required
    def index():
        tab = request.args.get("tab", "registrar")
        rol = request.args.get("rol")
        fecha = request.args.get("fecha")  # YYYY-MM-DD (compat)
        desde = request.args.get("desde")  # YYYY-MM-DD
        hasta = request.args.get("hasta")  # YYYY-MM-DD
        grado_f = request.args.get("grado")
        seccion_f = request.args.get("seccion")
        qname = request.args.get("q")
        q = (request.args.get("q") or "").strip()
        solo_activos = request.args.get("solo_activos") in ("1", "true", "on", "True")
        try:
            page = max(1, int(request.args.get("page", "1")))
        except Exception:
            page = 1
        try:
            per_page = int(request.args.get("per_page", "15"))
        except Exception:
            per_page = 15
        per_page = 15 if per_page <= 0 else min(per_page, 100)

        # Personas list (with optional search + pagination)
        personas_query = Persona.query.order_by(Persona.id.asc())
        if q:
            like = f"%{q}%"
            personas_query = personas_query.filter(
                or_(
                    Persona.nombre.ilike(like),
                    Persona.documento.ilike(like),
                    Persona.grado.ilike(like),
                    Persona.seccion.ilike(like),
                )
            )
        if rol:
            personas_query = personas_query.filter(Persona.rol == rol)
        if solo_activos:
            personas_query = personas_query.filter(Persona.activo.is_(True))
        personas_total = personas_query.count()
        personas = personas_query.offset((page - 1) * per_page).limit(per_page).all()
        personas_pages = max(1, (personas_total + per_page - 1) // per_page)

        asist_query = Asistencia.query.order_by(Asistencia.id.desc())
        # Date filtering: prefer range if provided; fallback to single-day 'fecha'
        try:
            if desde or hasta:
                if desde:
                    d0 = datetime.strptime(desde, "%Y-%m-%d")
                else:
                    # very early date if only hasta provided
                    d0 = datetime(1970, 1, 1)
                if hasta:
                    h = datetime.strptime(hasta, "%Y-%m-%d")
                    d1 = datetime(h.year, h.month, h.day, 23, 59, 59)
                else:
                    # very late date if only desde provided
                    d1 = datetime(2999, 12, 31, 23, 59, 59)
                asist_query = asist_query.filter(Asistencia.timestamp.between(d0, d1))
            elif fecha:
                d0 = datetime.strptime(fecha, "%Y-%m-%d")
                d1 = datetime(d0.year, d0.month, d0.day, 23, 59, 59)
                asist_query = asist_query.filter(Asistencia.timestamp.between(d0, d1))
        except Exception:
            pass
        # Join persona to enable filters by rol, grado, seccion, nombre
        asist_query = asist_query.join(Persona)
        if rol:
            asist_query = asist_query.filter(Persona.rol == rol)
        if grado_f:
            asist_query = asist_query.filter(Persona.grado == grado_f)
        if seccion_f:
            asist_query = asist_query.filter(Persona.seccion == seccion_f)
        if qname:
            like = f"%{qname}%"
            asist_query = asist_query.filter(Persona.nombre.ilike(like))
        # For Toma tab, show fewer rows by default
        default_limit = 20 if tab == "toma" else 100
        asistencias = asist_query.limit(default_limit).all()

        serial_log = serial_manager.get_log(100) if serial_manager else []

        # Derived report rows when requested
        report_rows = []
        report_summary = {}
        if tab == "reportes":
            report_rows = _build_report_rows(asistencias, app.tz)
            # Build summary
            try:
                total = len(report_rows)
                tard_count = sum(1 for r in report_rows if r.get("tardanza_min", 0) and r.get("tardanza_min", 0) > 0)
                tard_min = sum(int(r.get("tardanza_min", 0) or 0) for r in report_rows)
                dur_min = sum(int(r.get("duracion_min", 0) or 0) for r in report_rows)
                prom_dur = int(dur_min / total) if total else 0
                personas_unicas = len({r["person_id"] for r in report_rows})
                report_summary = {
                    "total_registros": total,
                    "personas": personas_unicas,
                    "tardanzas": tard_count,
                    "min_tardanza": tard_min,
                    "min_duracion": dur_min,
                    "prom_duracion": prom_dur,
                }
            except Exception:
                report_summary = {}

        return render_template(
            "index.html",
            personas=personas,
            asistencias=asistencias,
            rol=rol or "",
            fecha=fecha or "",
            desde=desde or "",
            hasta=hasta or "",
            serial_log=serial_log,
            tab=tab,
            timezone_name=app.config.get("TIMEZONE", "UTC"),
            report_rows=report_rows,
            report_summary=report_summary,
            q=q,
            page=page,
            pages=personas_pages,
            total=personas_total,
            per_page=per_page,
            solo_activos=solo_activos,
            grado=grado_f or "",
            seccion=seccion_f or "",
        )

    # Friendly tab routes
    @app.get("/registrar")
    @login_required
    def tab_registrar():
        return redirect(url_for("index", tab="registrar"))

    @app.get("/personas")
    @login_required
    def tab_personas():
        return redirect(url_for("index", tab="personas"))

    @app.get("/asistencias")
    @login_required
    def tab_asistencias():
        return redirect(url_for("index", tab="asistencias"))

    @app.get("/toma")
    @login_required
    def tab_toma():
        return redirect(url_for("index", tab="toma"))

    @app.get("/mensajes")
    @login_required
    def tab_mensajes():
        return redirect(url_for("index", tab="mensajes"))

    @app.get("/reportes")
    @login_required
    def tab_reportes():
        return redirect(url_for("index", tab="reportes"))

    # Auth
    @app.route("/login", methods=["GET", "POST"])
    def login():
        if request.method == "POST":
            from werkzeug.security import check_password_hash
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            user = User.query.filter_by(username=username).first()
            if user and check_password_hash(user.password_hash, password):
                login_user(user)
                flash("Sesión iniciada", "success")
                return redirect(url_for("index"))
            flash("Usuario o contraseña inválidos", "error")
            return redirect(url_for("login"))
        return render_template("login.html", tab="login")

    @app.get("/logout")
    def logout():
        try:
            logout_user()
        except Exception:
            pass
        flash("Sesión cerrada", "success")
        return redirect(url_for("login"))

    @app.route("/users", methods=["GET", "POST"])
    @login_required
    def users():
        # Only admins can create users
        if not getattr(current_user, "is_admin", False):
            flash("No autorizado", "error")
            return redirect(url_for("index"))
        if request.method == "POST":
            from werkzeug.security import generate_password_hash
            uname = request.form.get("username", "").strip()
            pwd = request.form.get("password", "")
            is_admin = True if request.form.get("is_admin") in ("1", "true", "on", "True") else False
            if not uname or not pwd:
                flash("Usuario y contraseña requeridos", "error")
                return redirect(url_for("users"))
            if User.query.filter_by(username=uname).first():
                flash("Usuario ya existe", "error")
                return redirect(url_for("users"))
            db.session.add(User(username=uname, password_hash=generate_password_hash(pwd), is_admin=is_admin))
            db.session.commit()
            flash("Usuario creado", "success")
            return redirect(url_for("users"))
        users = User.query.order_by(User.id.asc()).all()
        return render_template("users.html", users=users, tab="users")

    @app.post("/persona/<int:pid>/actualizar")
    @login_required
    def actualizar_persona(pid: int):
        persona = db.session.get(Persona, pid)
        if not persona:
            flash("Persona no encontrada", "error")
            return redirect(url_for("index", tab="personas"))

        nombre = request.form.get("nombre", "").strip()
        rol = request.form.get("rol", "").strip().upper()
        grado = request.form.get("grado", "").strip() or None
        seccion = request.form.get("seccion", "").strip() or None
        documento = request.form.get("documento", "").strip() or None
        activo = True if request.form.get("activo") in ("1", "true", "on", "True") else False

        if rol and rol not in ("ALUMNO", "PROFESOR", "PERSONAL"):
            flash("Rol inválido", "error")
            return redirect(url_for("index", tab="personas"))
        if documento:
            dup = Persona.query.filter(Persona.documento == documento, Persona.id != pid).first()
            if dup is not None:
                flash("Documento ya registrado en otra persona", "error")
                return redirect(url_for("index", tab="personas"))

        if nombre:
            persona.nombre = nombre
        if rol:
            persona.rol = rol
        persona.grado = grado
        persona.seccion = seccion
        persona.documento = documento
        persona.activo = activo
        db.session.commit()
        flash(f"Persona ID {pid} actualizada", "success")
        return redirect(url_for("index", tab="personas"))

    @app.post("/persona/<int:pid>/toggle")
    @login_required
    def toggle_persona(pid: int):
        persona = db.session.get(Persona, pid)
        if not persona:
            flash("Persona no encontrada", "error")
            return redirect(url_for("index", tab="personas"))
        persona.activo = not bool(persona.activo)
        db.session.commit()
        estado = "activada" if persona.activo else "desactivada"
        flash(f"Persona ID {pid} {estado}", "success")
        return redirect(url_for("index", tab="personas"))

    def _wants_json():
        best = request.accept_mimetypes.best_match(["application/json", "text/html"]) or "text/html"
        return best == "application/json" and (request.accept_mimetypes[best] > request.accept_mimetypes["text/html"])

    @app.post("/registrar")
    @login_required
    def registrar():
        nombre = request.form.get("nombre", "").strip()
        rol = request.form.get("rol", "").strip().upper()
        grado = request.form.get("grado", "").strip() or None
        seccion = request.form.get("seccion", "").strip() or None
        documento = request.form.get("documento", "").strip() or None
        if rol not in ("ALUMNO", "PROFESOR", "PERSONAL"):
            if _wants_json():
                return jsonify({"error": "Rol invalido"}), 400
            flash("Rol inválido", "error")
            return redirect(url_for("index"))
        if not nombre:
            if _wants_json():
                return jsonify({"error": "Nombre requerido"}), 400
            flash("Nombre requerido", "error")
            return redirect(url_for("index"))

        # Documento duplicado
        if documento:
            dup = Persona.query.filter_by(documento=documento).first()
            if dup is not None:
                if _wants_json():
                    return jsonify({"error": "Documento ya registrado"}), 400
                flash("Documento ya registrado en otra persona", "error")
                return redirect(url_for("index", tab="registrar"))

        # Find next available sensor ID (smallest positive int not used)
        used = {pid for (pid,) in db.session.query(Persona.id).all()}
        fid = 1
        while fid in used:
            fid += 1

        persona = Persona(id=fid, nombre=nombre, rol=rol, huella_guardada=False,
                          grado=grado, seccion=seccion, documento=documento)
        db.session.add(persona)
        db.session.commit()

        # Send serial command R,<ID> and wait for REG_OK/REG_FAIL
        try:
            serial_manager.send_cmd(f"R,{fid}")
        except Exception as e:
            # Rollback persona if serial not available
            db.session.delete(persona)
            db.session.commit()
            if _wants_json():
                return jsonify({"error": f"Serial no disponible: {e}"}), 500
            flash(f"Serial no disponible: {e}", "error")
            return redirect(url_for("index"))

        result = serial_manager.wait_reg_result(fid, app.config.get("REG_TIMEOUT", 30.0))
        if result == "REG_OK":
            # éxito: marcar huella
            persona.huella_guardada = True
            db.session.commit()
            if _wants_json():
                return jsonify({"status": "ok", "id": fid})
            flash(f"Registro exitoso de {nombre} (ID {fid}, rol {rol})", "success")
            return redirect(url_for("index", tab="personas"))
        else:
            # fallo o timeout: eliminar fila creada y no volver a tocar 'persona'
            pid_tmp = persona.id
            try:
                db.session.delete(persona)
                db.session.commit()
            finally:
                persona = None  # evitar accesos posteriores
            if _wants_json():
                return jsonify({"status": "fail", "reason": "timeout" if result is None else "reg_fail"}), 408 if result is None else 400
            motivo = "Tiempo agotado" if result is None else "Fallo de registro"
            flash(f"Registro fallido (ID {pid_tmp}): {motivo}", "error")
            return redirect(url_for("index", tab="registrar"))

    @app.post("/eliminar/<int:pid>")
    @login_required
    def eliminar(pid: int):
        persona = db.session.get(Persona, pid)
        if persona is None:
            if _wants_json():
                return jsonify({"error": "Persona no encontrada"}), 404
            flash("Persona no encontrada", "error")
            return redirect(url_for("index"))
        # If no fingerprint stored, just update state
        if not persona.huella_guardada:
            flash("La persona no tiene huella almacenada", "warning")
            return redirect(url_for("index"))

        try:
            serial_manager.send_cmd(f"E,{pid}")
        except Exception as e:
            if _wants_json():
                return jsonify({"error": f"Serial no disponible: {e}"}), 500
            flash(f"Serial no disponible: {e}", "error")
            return redirect(url_for("index"))

        result = serial_manager.wait_del_result(pid, app.config.get("REG_TIMEOUT", 30.0))
        if result == "DEL_OK":
            persona.huella_guardada = False
            db.session.commit()
            if not _wants_json():
                flash(f"Huella eliminada (ID {pid})", "success")
                return redirect(url_for("index"))
            return jsonify({"status": "ok", "id": pid})
        else:
            # keep state unchanged on failure/timeout
            if _wants_json():
                return jsonify({"status": "fail", "reason": "timeout" if result is None else "del_fail"}), 408 if result is None else 400
            motivo = "Tiempo agotado" if result is None else "Fallo al eliminar"
            flash(f"No se pudo eliminar la huella (ID {pid}): {motivo}", "error")
            return redirect(url_for("index"))

    @app.get("/health")
    @login_required
    def health():
        ports = []
        try:
            if serial is not None:
                ports = [p.device for p in serial.tools.list_ports.comports()]
        except Exception:
            ports = []
        return jsonify({
            "serial_config": {
                "port": app.config.get("SERIAL_PORT"),
                "baudrate": app.config.get("SERIAL_BAUDRATE"),
            },
            "serial_open": bool(serial_manager and serial_manager._ser and serial_manager._ser.is_open),
            "serial_in_use": getattr(serial_manager, "_current_port", None),
            "serial_enabled": bool(str(app.config.get("SERIAL_PORT", "")).strip().upper() not in {"", "OFF", "NONE", "DISABLE", "DISABLED"}),
            "available_ports": ports,
            "db_ok": True,
        })

    @app.get("/serial")
    @login_required
    def serial_log_api():
        try:
            limit = int(request.args.get("limit", "100"))
        except Exception:
            limit = 100
        lines = serial_manager.get_log(limit) if serial_manager else []
        return jsonify({
            "log": lines,
            "serial_open": bool(serial_manager and serial_manager._ser and serial_manager._ser.is_open),
        })

    @app.post("/persona/<int:pid>/borrar")
    @login_required
    def borrar_persona(pid: int):
        persona = db.session.get(Persona, pid)
        if persona is None:
            if _wants_json():
                return jsonify({"error": "Persona no encontrada"}), 404
            flash("Persona no encontrada", "error")
            return redirect(url_for("index"))

        force = request.form.get("force", "0") in ("1", "true", "True")

        if persona.huella_guardada:
            try:
                serial_manager.send_cmd(f"E,{pid}")
            except Exception as e:
                if not force:
                    if _wants_json():
                        return jsonify({"error": f"Serial no disponible: {e}"}), 500
                    flash(f"Serial no disponible: {e}", "error")
                    return redirect(url_for("index"))
            else:
                result = serial_manager.wait_del_result(pid, app.config.get("REG_TIMEOUT", 30.0))
                if result != "DEL_OK" and not force:
                    if _wants_json():
                        return jsonify({"status": "fail", "reason": "timeout" if result is None else "del_fail"}), 408 if result is None else 400
                    motivo = "Tiempo agotado" if result is None else "Fallo al eliminar"
                    flash(f"No se pudo eliminar la huella (ID {pid}): {motivo}", "error")
                    return redirect(url_for("index"))

        # Delete asistencias and persona
        try:
            Asistencia.query.filter_by(person_id=pid).delete()
            db.session.delete(persona)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            if _wants_json():
                return jsonify({"error": f"No se pudo borrar: {e}"}), 500
            flash(f"No se pudo borrar persona ID {pid}: {e}", "error")
            return redirect(url_for("index"))

        if not _wants_json():
            flash(f"Persona ID {pid} eliminada (y sus asistencias)", "success")
            return redirect(url_for("index"))
        return jsonify({"status": "ok", "id": pid})

    # Optional JSON endpoints
    @app.get("/api/personas")
    @login_required
    def api_personas():
        personas = Persona.query.order_by(Persona.id.asc()).all()
        return jsonify([p.as_dict() for p in personas])

    @app.get("/api/asistencias")
    @login_required
    def api_asistencias():
        try:
            limit = int(request.args.get("limit", "100"))
        except Exception:
            limit = 100

        rol = request.args.get("rol")
        fecha = request.args.get("fecha")  # YYYY-MM-DD
        desde = request.args.get("desde")  # YYYY-MM-DD
        hasta = request.args.get("hasta")  # YYYY-MM-DD

        asist_query = Asistencia.query.order_by(Asistencia.id.desc())
        try:
            if desde or hasta or fecha:
                if fecha and not (desde or hasta):
                    desde = fecha
                    hasta = fecha
                if desde:
                    d0_local = datetime.strptime(desde, "%Y-%m-%d")
                else:
                    d0_local = datetime(1970, 1, 1)
                if hasta:
                    h = datetime.strptime(hasta, "%Y-%m-%d")
                    d1_local = datetime(h.year, h.month, h.day, 23, 59, 59)
                else:
                    d1_local = datetime(2999, 12, 31, 23, 59, 59)
                # If timestamps are stored in UTC, convert local range to UTC; otherwise use local naive
                if not app.config.get("TIME_USE_LOCAL", True):
                    d0 = d0_local.replace(tzinfo=app.tz).astimezone(timezone.utc).replace(tzinfo=None)
                    d1 = d1_local.replace(tzinfo=app.tz).astimezone(timezone.utc).replace(tzinfo=None)
                else:
                    d0 = d0_local
                    d1 = d1_local
                asist_query = asist_query.filter(Asistencia.timestamp.between(d0, d1))
        except Exception:
            pass
        if rol:
            asist_query = asist_query.join(Persona).filter(Persona.rol == rol)

        asistencias = asist_query.limit(limit).all()
        out = []
        for a in asistencias:
            # Build UTC timestamp depending on storage mode
            if app.config.get("TIME_USE_LOCAL", True):
                # Stored as local naive
                dt_local = a.timestamp.replace(tzinfo=app.tz)
                dt = dt_local.astimezone(timezone.utc)
            else:
                # Stored as UTC naive
                dt = a.timestamp.replace(tzinfo=timezone.utc)
            out.append({
                "id": a.id,
                "person_id": a.person_id,
                "timestamp": a.timestamp.isoformat(),
                "timestamp_utc": dt.isoformat().replace("+00:00", "Z"),
                "persona": {
                    "id": a.persona.id if a.persona else None,
                    "nombre": a.persona.nombre if a.persona else None,
                    "rol": a.persona.rol if a.persona else None,
                },
            })
        return jsonify(out)

    @app.get("/api/reportes.csv")
    @login_required
    def api_reportes_csv():
        rol = request.args.get("rol")
        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        grado_f = request.args.get("grado")
        seccion_f = request.args.get("seccion")
        qname = request.args.get("q")
        persona_id = request.args.get("persona_id")

        asist_query = Asistencia.query.order_by(Asistencia.timestamp.asc()).join(Persona)
        try:
            if desde or hasta:
                if desde:
                    d0 = datetime.strptime(desde, "%Y-%m-%d")
                else:
                    d0 = datetime(1970, 1, 1)
                if hasta:
                    h = datetime.strptime(hasta, "%Y-%m-%d")
                    d1 = datetime(h.year, h.month, h.day, 23, 59, 59)
                else:
                    d1 = datetime(2999, 12, 31, 23, 59, 59)
                # Convert range if storing UTC
                if not app.config.get("TIME_USE_LOCAL", True):
                    d0 = d0.replace(tzinfo=app.tz).astimezone(timezone.utc).replace(tzinfo=None)
                    d1 = d1.replace(tzinfo=app.tz).astimezone(timezone.utc).replace(tzinfo=None)
                asist_query = asist_query.filter(Asistencia.timestamp.between(d0, d1))
        except Exception:
            pass
        if rol:
            asist_query = asist_query.filter(Persona.rol == rol)
        if grado_f:
            asist_query = asist_query.filter(Persona.grado == grado_f)
        if seccion_f:
            asist_query = asist_query.filter(Persona.seccion == seccion_f)
        if persona_id:
            try:
                pid = int(persona_id)
                asist_query = asist_query.filter(Asistencia.person_id == pid)
            except Exception:
                pass
        if qname:
            like = f"%{qname}%"
            asist_query = asist_query.filter(Persona.nombre.ilike(like))

        asistencias = asist_query.all()
        rows = _build_report_rows(asistencias, app.tz)
        # Summary
        try:
            total = len(rows)
            tard_count = sum(1 for r in rows if r.get("tardanza_min", 0) and r.get("tardanza_min", 0) > 0)
            tard_min = sum(int(r.get("tardanza_min", 0) or 0) for r in rows)
            dur_min = sum(int(r.get("duracion_min", 0) or 0) for r in rows)
            prom_dur = int(dur_min / total) if total else 0
        except Exception:
            total = tard_count = tard_min = dur_min = prom_dur = 0

        sio = StringIO()
        delim = request.args.get("delim") or app.config.get("CSV_DELIMITER", ",")
        # Excel hint for delimiter + BOM for UTF-8
        try:
            sio.write("\ufeff")  # BOM
            sio.write(f"sep={delim}\n")
        except Exception:
            pass
        w = csv.writer(sio, delimiter=delim)
        w.writerow(["fecha", "person_id", "nombre", "rol", "grado", "seccion", "entrada", "salida", "tardanza_min", "duracion_min"])
        for r in rows:
            w.writerow([
                r["fecha"], r["person_id"], r["nombre"], r["rol"], r.get("grado"), r.get("seccion"), r["entrada"], r["salida"], r.get("tardanza_min", 0), r["duracion_min"],
            ])
        # Add summary at the end
        w.writerow([])
        w.writerow(["Resumen"])
        w.writerow(["total_registros", total])
        w.writerow(["tardanzas", tard_count])
        w.writerow(["min_tardanza", tard_min])
        w.writerow(["min_duracion", dur_min])
        w.writerow(["prom_duracion", prom_dur])
        resp = make_response(sio.getvalue())
        resp.headers["Content-Type"] = "text/csv; charset=utf-8"
        fname_date = datetime.now().strftime("%Y%m%d")
        resp.headers["Content-Disposition"] = f"attachment; filename=reporte_{fname_date}.csv"
        return resp

    @app.get("/api/reportes.xlsx")
    @login_required
    def api_reportes_xlsx():
        try:
            import openpyxl
            from openpyxl.styles import Font
        except Exception:
            return jsonify({"error": "openpyxl no instalado"}), 501

        rol = request.args.get("rol")
        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        grado_f = request.args.get("grado")
        seccion_f = request.args.get("seccion")
        qname = request.args.get("q")
        persona_id = request.args.get("persona_id")

        asist_query = Asistencia.query.order_by(Asistencia.timestamp.asc()).join(Persona)
        try:
            if desde or hasta:
                if desde:
                    d0 = datetime.strptime(desde, "%Y-%m-%d")
                else:
                    d0 = datetime(1970, 1, 1)
                if hasta:
                    h = datetime.strptime(hasta, "%Y-%m-%d")
                    d1 = datetime(h.year, h.month, h.day, 23, 59, 59)
                else:
                    d1 = datetime(2999, 12, 31, 23, 59, 59)
                if not app.config.get("TIME_USE_LOCAL", True):
                    d0 = d0.replace(tzinfo=app.tz).astimezone(timezone.utc).replace(tzinfo=None)
                    d1 = d1.replace(tzinfo=app.tz).astimezone(timezone.utc).replace(tzinfo=None)
                asist_query = asist_query.filter(Asistencia.timestamp.between(d0, d1))
        except Exception:
            pass
        if rol:
            asist_query = asist_query.filter(Persona.rol == rol)
        if grado_f:
            asist_query = asist_query.filter(Persona.grado == grado_f)
        if seccion_f:
            asist_query = asist_query.filter(Persona.seccion == seccion_f)
        if persona_id:
            try:
                pid = int(persona_id)
                asist_query = asist_query.filter(Asistencia.person_id == pid)
            except Exception:
                pass
        if qname:
            like = f"%{qname}%"
            asist_query = asist_query.filter(Persona.nombre.ilike(like))

        asistencias = asist_query.all()
        rows = _build_report_rows(asistencias, app.tz)
        # Summary
        try:
            total = len(rows)
            tard_count = sum(1 for r in rows if r.get("tardanza_min", 0) and r.get("tardanza_min", 0) > 0)
            tard_min = sum(int(r.get("tardanza_min", 0) or 0) for r in rows)
            dur_min = sum(int(r.get("duracion_min", 0) or 0) for r in rows)
            prom_dur = int(dur_min / total) if total else 0
        except Exception:
            total = tard_count = tard_min = dur_min = prom_dur = 0

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reportes"
        # Summary sheet
        ws_sum = wb.create_sheet("Resumen", 0)
        ws_sum.append(["Métrica", "Valor"])
        for c in ws_sum[1]:
            c.font = Font(bold=True)
        for k, v in (
            ("Total registros", total),
            ("Personas", len({r["person_id"] for r in rows})),
            ("Tardanzas", tard_count),
            ("Minutos de tardanza", tard_min),
            ("Minutos de duración", dur_min),
            ("Promedio duración (min)", prom_dur),
        ):
            ws_sum.append([k, v])
        ws_sum.column_dimensions['A'].width = 28
        ws_sum.column_dimensions['B'].width = 18
        ws_sum.freeze_panes = "A2"
        headers = ["fecha", "person_id", "nombre", "rol", "grado", "seccion", "entrada", "salida", "tardanza_min", "duracion_min"]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in rows:
            ws.append([
                r["fecha"], r["person_id"], r["nombre"], r["rol"], r.get("grado"), r.get("seccion"), r["entrada"], r["salida"], r.get("tardanza_min", 0), r["duracion_min"],
            ])
        # Ajuste básico de anchos
        for col in ws.columns:
            maxlen = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > maxlen:
                    maxlen = len(val)
            ws.column_dimensions[col_letter].width = min(maxlen + 2, 40)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        from io import BytesIO
        bio = BytesIO()
        wb.save(bio)
        resp = make_response(bio.getvalue())
        resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        fname_date = datetime.now().strftime("%Y%m%d")
        resp.headers["Content-Disposition"] = f"attachment; filename=reporte_{fname_date}.xlsx"
        return resp

    @app.get("/api/personas.csv")
    @login_required
    def api_personas_csv():
        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        q = (request.args.get("q") or "").strip()
        rol = (request.args.get("rol") or "").strip().upper()
        solo_activos = request.args.get("solo_activos") in ("1", "true", "on", "True")

        personas_query = Persona.query.order_by(Persona.id.asc())
        if q:
            like = f"%{q}%"
            personas_query = personas_query.filter(
                or_(
                    Persona.nombre.ilike(like),
                    Persona.documento.ilike(like),
                    Persona.grado.ilike(like),
                    Persona.seccion.ilike(like),
                )
            )
        if rol:
            personas_query = personas_query.filter(Persona.rol == rol)
        if solo_activos:
            personas_query = personas_query.filter(Persona.activo.is_(True))

        personas = personas_query.all()

        # Build date filters
        d0 = None
        d1 = None
        try:
            if desde:
                d0 = datetime.strptime(desde, "%Y-%m-%d")
            if hasta:
                h = datetime.strptime(hasta, "%Y-%m-%d")
                d1 = datetime(h.year, h.month, h.day, 23, 59, 59)
        except Exception:
            d0 = d0

        sio = StringIO()
        delim = request.args.get("delim") or app.config.get("CSV_DELIMITER", ",")
        try:
            sio.write("\ufeff"); sio.write(f"sep={delim}\n")
        except Exception:
            pass
        w = csv.writer(sio, delimiter=delim)
        w.writerow(["id", "nombre", "rol", "grado", "seccion", "documento", "activo", "huella_guardada", "asistencias_count"])
        for p in personas:
            q = Asistencia.query.filter_by(person_id=p.id)
            if d0 and d1:
                q = q.filter(Asistencia.timestamp.between(d0, d1))
            elif d0:
                q = q.filter(Asistencia.timestamp >= d0)
            elif d1:
                q = q.filter(Asistencia.timestamp <= d1)
            count = q.count()
            w.writerow([p.id, p.nombre, p.rol, p.grado or "", p.seccion or "", p.documento or "", int(p.activo), int(p.huella_guardada), count])

        resp = make_response(sio.getvalue())
        resp.headers["Content-Type"] = "text/csv; charset=utf-8"
        fname_date = datetime.now().strftime("%Y%m%d")
        resp.headers["Content-Disposition"] = f"attachment; filename=personas_{fname_date}.csv"
        return resp

    @app.get("/api/asistencias.csv")
    @login_required
    def api_asistencias_csv():
        rol = request.args.get("rol")
        fecha = request.args.get("fecha")  # YYYY-MM-DD
        desde = request.args.get("desde")  # YYYY-MM-DD
        hasta = request.args.get("hasta")  # YYYY-MM-DD
        try:
            limit = request.args.get("limit")
            limit = int(limit) if limit is not None else None
        except Exception:
            limit = None

        asist_query = Asistencia.query.order_by(Asistencia.id.desc())
        try:
            if desde or hasta or fecha:
                if fecha and not (desde or hasta):
                    desde = fecha
                    hasta = fecha
                if desde:
                    d0_local = datetime.strptime(desde, "%Y-%m-%d")
                else:
                    d0_local = datetime(1970, 1, 1)
                if hasta:
                    h = datetime.strptime(hasta, "%Y-%m-%d")
                    d1_local = datetime(h.year, h.month, h.day, 23, 59, 59)
                else:
                    d1_local = datetime(2999, 12, 31, 23, 59, 59)
                if not app.config.get("TIME_USE_LOCAL", True):
                    d0 = d0_local.replace(tzinfo=app.tz).astimezone(timezone.utc).replace(tzinfo=None)
                    d1 = d1_local.replace(tzinfo=app.tz).astimezone(timezone.utc).replace(tzinfo=None)
                else:
                    d0 = d0_local
                    d1 = d1_local
                asist_query = asist_query.filter(Asistencia.timestamp.between(d0, d1))
        except Exception:
            pass
        if rol:
            asist_query = asist_query.join(Persona).filter(Persona.rol == rol)

        if limit:
            asistencias = asist_query.limit(limit).all()
        else:
            asistencias = asist_query.all()

        sio = StringIO()
        delim = request.args.get("delim") or app.config.get("CSV_DELIMITER", ",")
        try:
            sio.write("\ufeff"); sio.write(f"sep={delim}\n")
        except Exception:
            pass
        w = csv.writer(sio, delimiter=delim)
        w.writerow(["id", "person_id", "persona_nombre", "persona_rol", "timestamp"])
        for a in asistencias:
            nombre = a.persona.nombre if a.persona else None
            prole = a.persona.rol if a.persona else None
            w.writerow([a.id, a.person_id, nombre, prole, a.timestamp.isoformat()])

        resp = make_response(sio.getvalue())
        resp.headers["Content-Type"] = "text/csv; charset=utf-8"
        fname_date = datetime.now().strftime("%Y%m%d")
        resp.headers["Content-Disposition"] = f"attachment; filename=asistencias_{fname_date}.csv"
        return resp

    # Start serial worker once (can be disabled via SERIAL_PORT)
    global serial_manager
    # In debug, only start the worker in the reloader child process
    if os.environ.get("WERKZEUG_RUN_MAIN") == "true" or not app.debug:
        port_cfg = str(app.config.get("SERIAL_PORT", "")).strip()
        if port_cfg and port_cfg.upper() not in {"OFF", "NONE", "DISABLE", "DISABLED"}:
            # Avoid double thread on debug reload
            serial_manager = SerialManager(app)
            serial_manager.start()
        else:
            serial_manager = None

    return app


app = create_app()

if __name__ == "__main__":
    import os as _os
    _debug = (_os.getenv("FLASK_DEBUG", "0") == "1")
    # Run without reloader to avoid double-opening the serial port
    app.run(host="0.0.0.0", port=5000, debug=_debug, use_reloader=False)
